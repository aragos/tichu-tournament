#!/usr/bin/python

from calculator import Calls
from calculator import HandResult
from calculator import OrderBy
from calculator import Board
from calculator import Calculate
from calculator import TeamSummary
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors
from openpyxl.styles import fills
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook
import csv

def ReadXlsxInput(filename):
  """ Reads input from an Xlsx file. 
  
  Input file must have a sheet titled "Team Names" and sheet titled "Raw Hand
  Scores". This code will read the first continuous block of hand records, 
  stopping at the first blank line.
    
  Args:
    filename: path to the input file. 
    
  Returns:
    A tuple (input wb object, 
             list of Boards from the input in no defined order).
  """

  wb = load_workbook(filename)
  board_no_to_hr_list = {}
  teams_sheet = wb["Team Names"]
  raw_input_sheet = wb["Raw Hand Scores"]
  first = True
  for row in raw_input_sheet.rows:
    if not row[0].value:
      break
    if first:
      first = False
    else:
      board_no = int(row[0].value)
      hr_list = board_no_to_hr_list.setdefault(board_no, [])
      # For Tichu calls, need to make sure empty cells default to ''
      for i in xrange(3, 7):
        row[i].value = '' if not row[i].value else row[i].value
      hr_list.append(
        HandResult(board_no, int(row[1].value), int(row[2].value),
                   int(row[7].value), int(row[8].value),
                   Calls(row[3].value, row[4].value, row[5].value,
                         row[6].value)))
  board_list = []
  for k, v in board_no_to_hr_list.items():
    board_list.append(Board(k, v))  
  return (wb, board_list)


def SetNumberFormat(cell, format):
  cell.number_format = format


def SetAlignment(cell, alignment):
  cell.alignment = alignment


def SetFill(cell, color):
  cell.fill = PatternFill(fill_type=fills.FILL_SOLID, start_color=color, 
                          end_color=color)


def SetFont(cell, bold, color=colors.BLACK):
  cell.font = Font(bold=bold, color=color)


def SetBorder(cell, left=False, right=False, top=False, bottom=False):
  left_border = 'thin' if left else 'none'
  right_border = 'thin' if right else 'none'
  top_border = 'thin' if top else 'none'
  bottom_border = 'thin' if bottom  else 'none'
  cell.border = Border(left=Side(style=left_border), 
                       right=Side(style=right_border), 
                       top=Side(style=top_border), 
                       bottom=Side(style=bottom_border))


def SetColumnStyle(sheet, column, style_fun):
  cell_range = column + '1:' + column + '1000'
  for meta_cell in sheet[cell_range]:
    for cell in meta_cell:
      style_fun(cell)


def SetSheetHeaders(sheet):
  """ Sets the style of the top line in sheet with a frozen top row.
  
  Args:
    sheet: worksheet in which to set styles.
  """
  
  for meta_cell in sheet['A1:U1']:
    for cell in meta_cell:
      cell.font = Font(bold=True)
      cell.alignment = Alignment(horizontal='center')
  sheet.freeze_panes = sheet['A2']


SECTION_HEADER_COLOR = '6D9EEB'
def SetSectionHeaderStyleAndText(sheet, row_no, col_no, num_cols, text_list):
  """ Sets the style and header text for a section header.
  
  Args:
    sheet: worksheet in which to set style and text.
    row_no: row in which section header is set.
    col_no: first column of the header.
    num_cols: number of columns in the header.
    text_list: If text_list has only one member merges the header cells and
      centers the text in it. Otherwise the length of text_list must be the
      size of num_cols, and each column is populated with the corresponding text
      in text_list.
  """
  
  assert(len(text_list) == num_cols or len(text_list) == 1)
  for col in xrange(col_no, col_no + num_cols):
    cell = sheet.cell(row=row_no, column=col)
    SetFill(cell, SECTION_HEADER_COLOR)
    SetAlignment(cell, Alignment(horizontal='center'))
    SetFont(cell, True, color=colors.WHITE)

    if len(text_list) > 1:
      if col == col_no:
        SetBorder(cell, True, False, True, True)
      elif col == col_no + num_cols:
        SetBorder(cell, False, True, True, True)
      else:
        SetBorder(cell, False, False, True, True)
      cell.value = text_list[col-col_no]
  
  if len(text_list) == 1:
    cell = sheet.cell(row=row_no, column = col_no);
    cell.value = text_list[0]
    sheet.merge_cells(start_row=row_no, start_column=col_no, end_row=row_no,
                      end_column=num_cols)
    SetBorder(cell, True, True, True, True)
    

def SetDataTableStyle(sheet, start_row, start_col, num_rows, num_cols,
                      color=None):
  """ Sets the style and border around a data table.
  
  Args:
    sheet: worksheet in which to set style.
    start_row: row with the first set of data.
    start_col: column with the first set of data.
    num_rows: number of rows in the dataset.
    num_cols: number of columns in the datast.
    color: if set - all cells will be filled with color.
  """

  # Set the alignment to be 'right' everywhere but the headers.
  for row_no in xrange(start_row, start_row + num_rows):
    for col_no in xrange(start_col, num_cols + start_col):
      SetAlignment(sheet.cell(column=col_no, row=row_no),
                   Alignment(horizontal='right'))

  # Fill in right and left borders on all rows but the last one.
  for row in xrange(start_row, start_row + num_rows - 1):
    SetBorder(sheet.cell(row=row, column=start_col + num_cols - 1),
              False, True, False, False)
    SetBorder(sheet.cell(row=row, column=start_col), True, False, False, False)

  
  # Fill in bottom border on the last row for all columns but the first and
  # last.  
  for col in xrange(start_col + 1, num_cols + start_col - 1):
    SetBorder(sheet.cell(row=start_row + num_rows - 1, column=col),
        False, False, False, True)

  # Fill in broders of bottom corners.
  SetBorder(sheet.cell(row=start_row + num_rows - 1,
                       column=start_col + num_cols - 1),
            False, True, False, True)
  SetBorder(sheet.cell(row=start_row + num_rows - 1, column=start_col),
            True, False, False, True)

  if color:
    for row in xrange(start_row, start_row + num_rows):
      for col in xrange(start_col, num_cols + start_col):
        SetFill(sheet.cell(row=row, column=col), color)

SUMMARY_TABLE_OFFSET = 3
SUMMARY_TABLE_COLOR = 'FCEAD8'
RANK_TEXT = "Rank"
TEAM_TEXT = "Team" 
BOARD_NO_TEXT = "Board No"
MPS_TEXT = "MPs"
RPS_TEXT = "RPs"
APS_TEXT = "APs"
SIT_OUT_BONUS_TEXT = "Sit-out Bonus"
  
def WriteXlsxAggressivenessSummaries(max_rounds, scores, sheet):
  """ Writes a summary of play by aggressiveness to sheet.
 
  Two main components of the sheet. Starting at column 1, there will be a list
  of data tables for each team, with the aggressiveness points for each
  board in order of descending aggressiveness. To the right on top of the page
  will be a list of all teams ranked by total aggressiveness.
 
  
  Args:
    max_rounds: Max number of rounds expected for each team. Teams that played 
        fewer rounds will have an extra field with a bonus sit-out offset.
    scores: TeamSummaries listed in decreasing order of aggressiveness.
    sheet: sheet in which the summaries will be written.
  """

  sheet.title = 'Summary by Agressiveness'

  headers = [BOARD_NO_TEXT, APS_TEXT]
  sheet.append(headers)
  SetSheetHeaders(sheet)
  
  # Write a table for each team represented in scores.
  row_no = 1
  for s in scores:
    row_no += 1
    SetSectionHeaderStyleAndText(sheet, row_no, 1, len(headers),
                                 ["Team {0}".format(s.team_no)])
    start_row = row_no + 1
    for key in sorted(s.board_aps.keys()):
      row_no += 1
      row_dict = [key, s.board_aps[key]]
      for col_no in range(1, len(headers) + 1):
        sheet.cell(column=col_no, row=row_no, value = row_dict[col_no - 1])
    # This is kind of a hack, but back calculate how much was added from the 
    # sit-out bonus.
    if len(s.board_aps) < max_rounds:
      row_no += 1
      row_dict = [SIT_OUT_BONUS_TEXT, 
                  s.aps - s.aps * len(s.board_aps) / max_rounds]
      for col_no in range(1, len(headers) + 1):
        sheet.cell(column=col_no, row=row_no, value = row_dict[col_no - 1])
        SetAlignment(sheet.cell(column=col_no, row=row_no),
                     Alignment(horizontal='right'))
    
    SetDataTableStyle(sheet, start_row, 1, 
                      len(s.board_aps) + 1 if len(s.board_aps) < max_rounds else len(s.board_aps),
                      len(headers))
    # Space before the next set of scores.
    sheet.append([])
    row_no += 1
  
  # Write a summary table with just the totals.
  summary_header = [RANK_TEXT, TEAM_TEXT, APS_TEXT]
  SetSectionHeaderStyleAndText(sheet, 2, len(headers) + SUMMARY_TABLE_OFFSET,
                               len(summary_header), summary_header)
  for i in range(len(scores)):
    sheet.cell(column=len(headers) + SUMMARY_TABLE_OFFSET, row=3 + i,
       value = i + 1)
    sheet.cell(column=len(headers) + SUMMARY_TABLE_OFFSET + 1, row=3 + i,
       value = scores[i].team_no)
    sheet.cell(column=len(headers) + SUMMARY_TABLE_OFFSET + 2, row=3 + i,
       value = scores[i].aps)
  SetDataTableStyle(sheet, 3, len(headers) + SUMMARY_TABLE_OFFSET, len(scores),
                    len(summary_header), SUMMARY_TABLE_COLOR)

  # Stylistic things to make floats fit into columns.
  SetColumnStyle(sheet, 'A', lambda x: SetNumberFormat(x, '0')) 
  SetColumnStyle(sheet, 'B', lambda x: SetNumberFormat(x, '0.0')) 
  SetColumnStyle(sheet, 'G', lambda x: SetNumberFormat(x, '0.0')) 
  sheet.column_dimensions['A'].width = 13
  sheet.column_dimensions['B'].width = 10 
  sheet.row_dimensions[1].height = 15


def WriteXlsxTeamSummaries(max_rounds, scores, sheet):
  """ Writes a summary of play for each team to sheet.
 
  Two main components of the sheet. Starting at column 1, there will be a list
  of data tables for each team, with the MPs and RPs for each board in order of
  descending MPs. To the right on top of the page will be a list of all teams
  ranked by total MPs.
  
  Args:
    max_rounds: Max number of rounds expected for each team. Teams that played 
        fewer rounds will have an extra field with a bonus sit-out offset.
    scores: TeamSummaries listed in decreasing order of aggressiveness.
    sheet: sheet in which the summaries will be written.
  """

  sheet.title = 'Summary by Team'

  headers = [BOARD_NO_TEXT, MPS_TEXT, RPS_TEXT]
  sheet.append(headers)
  SetSheetHeaders(sheet)  

  # Write a table for each team represented in scores.
  row_no = 1
  for s in scores:
    row_no += 1
    SetSectionHeaderStyleAndText(
        sheet, row_no, 1, len(headers),
        ["Place {1}. Team {0}: MPs {2:.1f} RPs {3:.2f}".format(
            s.team_no, s.mp_rank, s.mps, s.rps)])
    start_row = row_no + 1
    keys = sorted(s.board_mps.keys())
    for key in keys:
      row_no+= 1
      row_dict = {BOARD_NO_TEXT: key, MPS_TEXT: s.board_mps[key],
                  RPS_TEXT: s.board_rps[key]}
      for col_no in xrange(1, len(headers) + 1):
        sheet.cell(column=col_no, row=row_no, value=row_dict[headers[col_no - 1]])

    if len(keys) < max_rounds:
      row_no += 1
      row_dict = {BOARD_NO_TEXT : SIT_OUT_BONUS_TEXT,
                  MPS_TEXT: s.mps - s.mps * len(keys) / max_rounds,
                  RPS_TEXT: s.rps - s.rps * len(keys) / max_rounds}
      for col_no in xrange(1, len(headers) + 1):
        sheet.cell(column=col_no, row=row_no,
                   value=row_dict[headers[col_no - 1]])
        SetAlignment(sheet.cell(column=col_no, row=row_no),
                     Alignment(horizontal='right'))
    SetDataTableStyle(sheet, start_row, 1,
                      len(keys) + 1 if len(keys) < max_rounds else len(keys),
                      len(headers))
    row_no += 1
    sheet.append([])
  
  # Write a summary table with just the total scores.
  summary_header = [RANK_TEXT, TEAM_TEXT, MPS_TEXT, RPS_TEXT]
  SetSectionHeaderStyleAndText(sheet, 2, len(headers) + SUMMARY_TABLE_OFFSET,
                               len(summary_header), summary_header)
  for i in range(len(scores)):
    sheet.cell(column=len(headers) + SUMMARY_TABLE_OFFSET, row=3 + i,
       value = i + 1)
    sheet.cell(column=len(headers) + SUMMARY_TABLE_OFFSET + 1, row=3 + i,
       value = scores[i].team_no)
    sheet.cell(column=len(headers) + SUMMARY_TABLE_OFFSET + 2, row=3 + i,
       value = scores[i].mps)
    sheet.cell(column=len(headers) + SUMMARY_TABLE_OFFSET + 3, row=3 + i,
       value = scores[i].rps)
  SetDataTableStyle(sheet, 3, len(headers) + SUMMARY_TABLE_OFFSET, len(scores),
                    len(summary_header), SUMMARY_TABLE_COLOR)

  # Stylistic things to make floats fit into columns.
  SetColumnStyle(sheet, 'H', lambda x: SetNumberFormat(x, '0.00')) 
  SetColumnStyle(sheet, 'I', lambda x: SetNumberFormat(x, '0.00')) 
  SetColumnStyle(sheet, 'C', lambda x: SetNumberFormat(x, '0.00'))
  SetColumnStyle(sheet, 'B', lambda x: SetNumberFormat(x, '0.0'))  
  SetColumnStyle(sheet, 'A', lambda x: SetNumberFormat(x, '0')) 
  sheet.column_dimensions['A'].width = 13 
  sheet.column_dimensions['B'].width = 10 
  sheet.column_dimensions['C'].width = 10 
  sheet.row_dimensions[1].height = 15


def WriteXlsxBoardSummaries(board_list, sheet):
  """ Writes a summary of play organized by board.
  
  Args:
    board_list: A list of Board objects in ascending order.
    sheet: sheet in which the summaries will be written.
  """

  sheet.title = 'Summary by Board'

  NS_TEAM_TEXT = 'NS Team'
  EW_TEAM_TEXT = 'EW Team'
  CALLS_TEXT = 'Calls'
  NS_SCORE_TEXT = 'NS Score'
  EW_SCORE_TEXT = 'EW Score'
  NS_MPS_TEXT = 'NS MPs'
  EW_MPS_TEXT = 'EW MPs'
  NS_RPS_TEXT = 'NS RPs'
  EW_RPS_TEXT = 'EW RPs'
  NS_APS_TEXT = 'NS APs'
  EW_APS_TEXT = 'EW APs'
  
  headers = [NS_TEAM_TEXT, EW_TEAM_TEXT, CALLS_TEXT, NS_SCORE_TEXT,
             EW_SCORE_TEXT, NS_MPS_TEXT, EW_MPS_TEXT, NS_RPS_TEXT, 
             EW_RPS_TEXT, NS_APS_TEXT, EW_APS_TEXT]
  sheet.append(headers)
  SetSheetHeaders(sheet)

  row_no = 1
  for b in board_list:
    row_no += 1
    SetSectionHeaderStyleAndText(sheet, row_no, 1, len(headers),
                                 ['Board {0}'.format(b._board_no)])
    start_row = row_no + 1
    for bs in b._board_score:
      row_no += 1
      row_dict = { NS_TEAM_TEXT: bs._hr.ns_pair_no(),
                   EW_TEAM_TEXT: bs._hr.ew_pair_no(),
                   CALLS_TEXT: str(bs._hr.calls()),
                   NS_SCORE_TEXT: bs._hr.ns_score(),
                   EW_SCORE_TEXT: bs._hr.ew_score(),
                   NS_MPS_TEXT: "{0:.1f}".format(bs.ns_mps),
                   EW_MPS_TEXT: "{0:.1f}".format(bs.ew_mps),
                   NS_RPS_TEXT: "{0:.2f}".format(bs.ns_rps),
                   EW_RPS_TEXT: "{0:.2f}".format(bs.ew_rps),
                   NS_APS_TEXT: "{0}".format(bs.ns_aps),
                   EW_APS_TEXT: "{0}".format(bs.ew_aps),}
      for col_no in range(1, len(headers)+1):
        sheet.cell(column=col_no, row=row_no,
                   value = row_dict[headers[col_no - 1]])	
    SetDataTableStyle(sheet, start_row, 1, len(b._board_score), len(headers))
    row_no += 1
    sheet.append([])

  sheet.column_dimensions['C'].width = 20 


def WriteHandNames(name_list, sheet):
  ''' Writes down players names for reference.
  Args:
    hand_list: List of pairs of hand in team numbr order. If a player name is
               unknown, must be None.
  '''
  sheet.append(["Team"])
  for i in range(len(name_list)):
    sheet.append([str(i + 1), name_list[i][0], name_list[i][1]])
  SetSheetHeaders(sheet)
  
  for row_no in xrange(2, len(name_list) + 2):
    cell = sheet.cell(row=row_no, column=1)
    SetAlignment(cell, Alignment(horizontal='right'))
  

def WriteXlsxRawScores(board_list, sheet):
  """ Writes down raw scores.
  
  Args:
    board_list: A list of Board objects in ascending order.
    sheet: sheet in which the summaries will be written.
  """

  sheet.title =  'Raw Hand Scores'

  BOARD_TEXT = 'Board'
  NS_TEAM_TEXT = 'NS Team'
  EW_TEAM_TEXT = 'EW Team'
  N_CALL = "N Call"
  S_CALL = "S Call"
  E_CALL = "E Call"
  W_CALL = "W Call"
  CALLS_TEXT = 'Calls'
  NS_SCORE_TEXT = 'NS Score'
  EW_SCORE_TEXT = 'EW Score'
  
  headers = [BOARD_TEXT, NS_TEAM_TEXT, EW_TEAM_TEXT, N_CALL, S_CALL,
             E_CALL, W_CALL, NS_SCORE_TEXT, EW_SCORE_TEXT]
  sheet.append(headers)
  SetSheetHeaders(sheet)

  row_no = 1
  for b in board_list:
    for bs in b._board_score:
      row_no += 1
      row_dict = { BOARD_TEXT: b._board_no,
                   NS_TEAM_TEXT: bs._hr.ns_pair_no(),
                   EW_TEAM_TEXT: bs._hr.ew_pair_no(),
                   N_CALL: bs._hr.calls().n_call(),
                   S_CALL: bs._hr.calls().s_call(),
                   E_CALL: bs._hr.calls().e_call(),
                   W_CALL: bs._hr.calls().w_call(),
                   NS_SCORE_TEXT: bs._hr.ns_score(),
                   EW_SCORE_TEXT: bs._hr.ew_score(),}
      for col_no in range(1, len(headers)+1):
          sheet.cell(column=col_no, row=row_no,
                     value = row_dict[headers[col_no - 1]])


def CopySheet(from_sheet, to_sheet):
  """ Copies all data from from_sheet to to_sheet. """

  for row in from_sheet.rows:
    row_list = []
    for cell in row:
      row_list.append(cell.value)
    to_sheet.append(row_list)


def WriteResultsToXlsx(max_rounds, mp_scores, ap_scores, board_list,
                       name_list=None, input_wb=None):
  """ Creates an Xlx workbook with all the information about a tournament.

      Args: 
        max_rounds: Max number of rounds expected for each team. Teams that 
          played fewer rounds have their scores adjusted to match the maximum 
          possible number of rounds played.
        mp_scores: TeamSummary objects ordered in descending total MP order.
        ap_scores: TeamSummary objects ordered in descending total AP order.
        board_list: List of boards in ascending board number order.
        name_list: List of player name pairs in ascending team number order.
          If no name exists for a player, must be None.
        input_wb: The input workbook as it was read in. Used to copy raw hand
          and team details into the output workbook. If None, raw scores and 
          team details will not be present in the output.
      Returns:
        formatted workbook
  """

  wb = Workbook()
  OrderBy(mp_scores, "MP")
  WriteXlsxTeamSummaries(max_rounds, mp_scores, wb.worksheets[0])
  board_sheet = wb.create_sheet()
  WriteXlsxBoardSummaries(board_list, board_sheet)
  aggro_sheet = wb.create_sheet()
  OrderBy(ap_scores, "AP")
  WriteXlsxAggressivenessSummaries(max_rounds, ap_scores, aggro_sheet)
  raw_scores_sheet = wb.create_sheet()
  WriteXlsxRawScores(board_list, raw_scores_sheet)
  # Copy the input sheets into the newly created workbook.
  TEAM_NAMES_TEXT = "Team Names"
  if name_list:
      name_sheet = wb.create_sheet(TEAM_NAMES_TEXT)
      WriteHandNames(name_list, name_sheet)
  elif input_wb:
      CopySheet(input_wb[TEAM_NAMES_TEXT], wb.create_sheet(TEAM_NAMES_TEXT))
      SetSheetHeaders(wb[TEAM_NAMES_TEXT])

  return wb


def OutputWorkbookAsBytesIO(wb):
  """ Writes the Xlx workbook to a BytesIO stream. """
  return BytesIO(save_virtual_workbook(wb))

